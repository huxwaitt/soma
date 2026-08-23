"""Attachment text extraction.

``extract_attachment_text`` saves one attachment to a temporary folder under
the user's profile, pulls the plain text out of it, and deletes the file
again. PDF and Excel need the optional ``search`` extra (``pypdf``,
``openpyxl``); Word, PowerPoint and plain text only need the standard
library.
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
import zipfile
from typing import Any
from xml.etree import ElementTree

from outlook_mcp.client.folders import _safe_get, get_item_by_id
from outlook_mcp.errors import OutlookError

KIND_BY_EXT = {
    ".pdf": "pdf",
    ".docx": "docx",
    ".xlsx": "xlsx",
    ".xlsm": "xlsx",
    ".pptx": "pptx",
    ".txt": "text",
    ".csv": "text",
    ".md": "text",
    ".log": "text",
    ".json": "text",
    ".xml": "text",
    ".html": "text",
    ".htm": "text",
}

_EXTRA_HINT = (
    "Install the 'search' extra of outlook-classic-mcp "
    "(uv sync --extra search, or pip install outlook-classic-mcp[search])."
)
_SAFE_EXT = re.compile(r"^\.[A-Za-z0-9]{1,10}$")

_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
_A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def kind_for(filename: str) -> str | None:
    ext = os.path.splitext(filename or "")[1].lower()
    return KIND_BY_EXT.get(ext)


def _temp_root() -> str:
    base = os.environ.get("LOCALAPPDATA") or os.environ.get("USERPROFILE") or tempfile.gettempdir()
    root = os.path.join(base, "outlook-mcp", "tmp")
    os.makedirs(root, exist_ok=True)
    return root


def _ooxml_paragraphs(path: str, member_pattern: str, para_tag: str, text_tag: str, tab_tag: str | None) -> str:
    """Concatenate paragraph text from the XML parts of an OOXML zip."""
    pattern = re.compile(member_pattern)
    out: list[str] = []
    with zipfile.ZipFile(path) as zf:
        members = [n for n in zf.namelist() if pattern.match(n)]

        def order(name: str) -> tuple[int, str]:
            m = re.search(r"(\d+)\.xml$", name)
            return (int(m.group(1)) if m else 0, name)

        for i, name in enumerate(sorted(members, key=order)):
            if i:
                out.append("")
            try:
                root = ElementTree.fromstring(zf.read(name))
            except ElementTree.ParseError:
                continue
            for para in root.iter(para_tag):
                pieces: list[str] = []
                for node in para.iter():
                    if node.tag == text_tag and node.text:
                        pieces.append(node.text)
                    elif tab_tag and node.tag == tab_tag:
                        pieces.append("\t")
                line = "".join(pieces)
                if line.strip():
                    out.append(line)
    return "\n".join(out)


def _extract_pdf(path: str) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise OutlookError("Reading PDF attachments needs pypdf. " + _EXTRA_HINT) from exc
    reader = PdfReader(path)
    pages: list[str] = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:  # noqa: BLE001 - one bad page should not lose the rest
            pages.append("")
    return "\n\n".join(p.strip() for p in pages).strip()


def _extract_docx(path: str) -> str:
    return _ooxml_paragraphs(path, r"^word/document\.xml$", _W_NS + "p", _W_NS + "t", _W_NS + "tab")


def _extract_pptx(path: str) -> str:
    return _ooxml_paragraphs(path, r"^ppt/slides/slide\d+\.xml$", _A_NS + "p", _A_NS + "t", None)


def _extract_xlsx(path: str) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise OutlookError("Reading Excel attachments needs openpyxl. " + _EXTRA_HINT) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[str] = []
        for ws in wb.worksheets:
            lines = [f"# {ws.title}"]
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v) for v in row]
                if any(c.strip() for c in cells):
                    lines.append("\t".join(cells).rstrip())
            sheets.append("\n".join(lines))
        return "\n\n".join(sheets)
    finally:
        wb.close()


def _extract_text(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        return fh.read()


_EXTRACTORS = {
    "pdf": _extract_pdf,
    "docx": _extract_docx,
    "xlsx": _extract_xlsx,
    "pptx": _extract_pptx,
    "text": _extract_text,
}


def extract_text_from_file(path: str, kind: str) -> str:
    """Plain text of a file of ``kind`` (pdf / docx / xlsx / pptx / text)."""
    try:
        extractor = _EXTRACTORS[kind]
    except KeyError:
        raise OutlookError(f"Unsupported attachment kind: {kind!r}.") from None
    try:
        text = extractor(path)
    except OutlookError:
        raise
    except zipfile.BadZipFile as exc:
        raise OutlookError("Attachment is not a valid Office file (zip container is damaged).") from exc
    except Exception as exc:  # noqa: BLE001
        raise OutlookError(f"Could not read attachment text: {exc}") from exc
    return text.replace("\r\n", "\n").replace("\r", "\n")


def extract_attachment_text(
    outlook: Any,
    namespace: Any,
    *,
    entry_id: str,
    index: int,
    max_chars: int = 20000,
) -> dict[str, Any]:
    item = get_item_by_id(namespace, entry_id)
    attachments = list(_safe_get(item, "Attachments") or [])
    if index < 1 or index > len(attachments):
        raise OutlookError(
            f"index {index} out of range (message has {len(attachments)} attachments, 1-indexed)."
        )
    att = attachments[index - 1]
    filename = str(_safe_get(att, "FileName", "") or "")
    kind = kind_for(filename)
    if kind is None:
        raise OutlookError(
            f"Cannot extract text from {filename!r}: supported types are "
            + ", ".join(sorted(KIND_BY_EXT))
            + "."
        )
    ext = os.path.splitext(filename)[1].lower()
    if not _SAFE_EXT.match(ext):
        raise OutlookError(f"Attachment extension {ext!r} rejected.")

    tmp_dir = tempfile.mkdtemp(prefix="att-", dir=_temp_root())
    tmp_path = os.path.join(tmp_dir, "attachment" + ext)
    try:
        att.SaveAsFile(tmp_path)
        text = extract_text_from_file(tmp_path, kind)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    chars = len(text)
    truncated = bool(max_chars) and chars > max_chars
    if truncated:
        text = text[:max_chars].rstrip()
    return {
        "entry_id": entry_id,
        "index": index,
        "filename": filename,
        "kind": kind,
        "chars": chars,
        "truncated": truncated,
        "text": text,
    }
